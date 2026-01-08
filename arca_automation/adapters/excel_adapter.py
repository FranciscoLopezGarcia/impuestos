"""
Excel Adapter - Actualiza hojas de Excel con parámetros ARCA.

Funciona con la estructura de tabla normalizada del Excel.
"""
from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config.settings import (
    EXCEL_SHEET_NAME,
    EXCEL_DATA_START_ROW,
    EXCEL_COLUMN_MAP,
    NUMERIC_COLUMNS,
    EXCEL_NUMBER_FORMAT,
)
from utils.logger import get_logger


class ExcelAdapter:
    """Actualiza hojas de Excel con parámetros ARCA."""
    
    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)
        self.logger = get_logger(self.__class__.__name__)
        
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel no encontrado: {excel_path}")
    
    def update(self, parametros: List[Dict], rebuild: bool = False):
        """
        Actualiza el Excel con los parámetros.
        
        Args:
            parametros: Lista de diccionarios con datos normalizados
            rebuild: Si True, borra todo y reescribe. Si False, actualiza existentes.
        """
        self.logger.info(f"Actualizando Excel: {self.excel_path.name}")
        self.logger.info(f"Registros a procesar: {len(parametros)}")
        self.logger.info(f"Modo: {'REBUILD' if rebuild else 'UPDATE'}")
        
        wb = load_workbook(self.excel_path)
        
        # Verificar que exista la hoja
        if EXCEL_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Hoja '{EXCEL_SHEET_NAME}' no encontrada en {self.excel_path.name}")
        
        ws = wb[EXCEL_SHEET_NAME]
        
        # Si rebuild, borrar datos existentes
        if rebuild:
            self._clear_data(ws)
        
        # Obtener mapeo de columnas
        headers = self._get_headers(ws)
        excel_cols = self._build_column_map(headers)
        
        # Obtener registros existentes (para UPDATE)
        existing = {}
        if not rebuild:
            existing = self._get_existing_records(ws, excel_cols)
        
        # Actualizar/insertar registros
        updated_count = 0
        inserted_count = 0
        
        for item in parametros:
            key = self._make_key(item)
            
            if key in existing:
                # Actualizar existente
                row = existing[key]
                self._write_row(ws, row, item, excel_cols)
                updated_count += 1
            else:
                # Insertar nuevo
                row = ws.max_row + 1
                self._write_row(ws, row, item, excel_cols)
                existing[key] = row
                inserted_count += 1
        
        # Guardar
        wb.save(self.excel_path)
        
        self.logger.info(f"✓ Excel actualizado:")
        self.logger.info(f"   Actualizados: {updated_count}")
        self.logger.info(f"   Insertados: {inserted_count}")
        self.logger.info(f"   Total: {updated_count + inserted_count}")
    
    def _clear_data(self, ws):
        """Borra todos los datos (conserva headers)."""
        if ws.max_row > 1:
            ws.delete_rows(EXCEL_DATA_START_ROW, ws.max_row - 1)
        self.logger.debug("Datos anteriores borrados")
    
    def _get_headers(self, ws) -> Dict[int, str]:
        """Obtiene headers de la primera fila."""
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[cell.column] = str(cell.value).strip()
        return headers
    
    def _build_column_map(self, headers: Dict[int, str]) -> Dict[str, int]:
        """
        Construye mapeo de campo → columna.
        
        Returns:
            Dict con {campo: columna_numero}
        """
        # Mapeo inverso: nombre_header → campo
        header_to_field = {
            "CONCEPTO": "concepto",
            "IMPUESTO": "impuesto",
            "AÑO": "anio",
            "VALOR": "valor",
            "DESDE": "desde",
            "HASTA": "hasta",
            "MONTO_FIJO": "monto_fijo",
            "PORCENTAJE": "porcentaje",
            "EXCEDENTE_DESDE": "excedente_desde",
            "UNIDAD": "unidad",
            "FUENTE": "fuente",
            "ORIGEN": "origen",
        }
        
        field_to_col = {}
        
        for col_num, header in headers.items():
            header_upper = header.upper()
            if header_upper in header_to_field:
                field = header_to_field[header_upper]
                field_to_col[field] = col_num
        
        return field_to_col
    
    def _get_existing_records(self, ws, excel_cols: Dict) -> Dict[tuple, int]:
        """
        Obtiene registros existentes.
        
        Returns:
            Dict con {(concepto, impuesto, anio): fila}
        """
        existing = {}
        
        concepto_col = excel_cols.get("concepto")
        impuesto_col = excel_cols.get("impuesto")
        anio_col = excel_cols.get("anio")
        
        if not all([concepto_col, impuesto_col, anio_col]):
            return existing
        
        for row in range(EXCEL_DATA_START_ROW, ws.max_row + 1):
            concepto = ws.cell(row, concepto_col).value
            impuesto = ws.cell(row, impuesto_col).value
            anio = ws.cell(row, anio_col).value
            
            if all([concepto, impuesto, anio]):
                key = (str(concepto), str(impuesto), int(anio))
                existing[key] = row
        
        return existing
    
    def _make_key(self, item: Dict) -> tuple:
        """Genera clave única para un registro."""
        return (
            str(item.get("concepto", "")),
            str(item.get("impuesto", "")),
            int(item.get("anio", 0))
        )
    
    def _write_row(self, ws, row: int, item: Dict, excel_cols: Dict):
        """Escribe un registro en una fila."""
        for field, value in item.items():
            col = excel_cols.get(field)
            if not col:
                continue
            
            cell = ws.cell(row, col)
            cell.value = value
            
            # Formatear números
            if field in NUMERIC_COLUMNS and isinstance(value, (int, float)):
                cell.number_format = EXCEL_NUMBER_FORMAT
    
    def validate_structure(self) -> bool:
        """
        Valida que el Excel tenga la estructura esperada.
        
        Returns:
            True si es válido, False si no
        """
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            
            if EXCEL_SHEET_NAME not in wb.sheetnames:
                self.logger.error(f"Hoja '{EXCEL_SHEET_NAME}' no encontrada")
                return False
            
            ws = wb[EXCEL_SHEET_NAME]
            headers = self._get_headers(ws)
            
            required_headers = {"CONCEPTO", "IMPUESTO", "AÑO"}
            found_headers = {h.upper() for h in headers.values()}
            
            if not required_headers.issubset(found_headers):
                missing = required_headers - found_headers
                self.logger.error(f"Headers faltantes: {missing}")
                return False
            
            wb.close()
            return True
        
        except Exception as e:
            self.logger.error(f"Error validando estructura: {e}")
            return False

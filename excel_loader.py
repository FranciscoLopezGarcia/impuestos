import json
from openpyxl import load_workbook
from pathlib import Path

# ==========================
# CONFIGURACIÓN
# ==========================

EXCEL_PATH = Path(r"C:\Users\franl\Desktop\impuestos\CH - Anuales 2024.xlsx")
JSON_PATH = Path(r"C:\Users\franl\Desktop\impuestos\outputs\parametros_arca.json")
SHEET_NAME = "Parametros_ARCA"

# 👉 PONER EN True SOLO ESTA VEZ
REBUILD = False

HEADER_MAP = {
    "concepto": "CONCEPTO",
    "impuesto": "IMPUESTO",
    "anio": "AÑO",
    "valor_num": "VALOR",
    "desde": "DESDE",
    "hasta": "HASTA",
    "monto_fijo": "MONTO_FIJO",
    "porcentaje": "PORCENTAJE",
    "excedente_desde": "EXCEDENTE_DESDE",
    "unidad": "UNIDAD",
    "fuente": "FUENTE",
    "origen": "ORIGEN",
}

NUMERIC_COLS = {"VALOR", "DESDE", "HASTA", "MONTO_FIJO", "EXCEDENTE_DESDE", "PORCENTAJE"}


# ==========================
# EJECUCIÓN
# ==========================

def main():
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 🧹 Limpieza controlada (solo una vez)
    if REBUILD:
        ws.delete_rows(2, ws.max_row)

    headers = [c.value for c in ws[1]]
    excel_cols = {h: i+1 for i, h in enumerate(headers)}

    existing = {}
    for r in range(2, ws.max_row + 1):
        key = (
            ws.cell(r, excel_cols["CONCEPTO"]).value,
            ws.cell(r, excel_cols["IMPUESTO"]).value,
            ws.cell(r, excel_cols["AÑO"]).value,
        )
        if all(key):
            existing[key] = r

    for item in data:
        key = (item["concepto"], item["impuesto"], item["anio"])

        if key in existing:
            row = existing[key]
        else:
            row = ws.max_row + 1
            existing[key] = row

        for j_key, x_key in HEADER_MAP.items():
            if j_key in item:
                cell = ws.cell(row, excel_cols[x_key])
                value = item[j_key]
                cell.value = value

                # 🧮 Blindaje contable
                if x_key in NUMERIC_COLS and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

    wb.save(EXCEL_PATH)
    print("✅ Excel actualizado correctamente (formato contable seguro)")


if __name__ == "__main__":
    main()

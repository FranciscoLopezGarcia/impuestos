import json
from openpyxl import load_workbook
from pathlib import Path
from config.paths import NORMALIZED_DIR

SHEET_NAME = "Parametros_ARCA"

HEADER_MAP = {
    "concepto": "CONCEPTO",
    "impuesto": "IMPUESTO",
    "anio": "AÑO",
    "valor_num": "VALOR",
    "desde": "DESDE",
    "hasta": "HASTA",
    "monto_fijo": "MONTO_FIJO",
    "porcentaje": "PORCENTAJE",
    "excedente_desde": "EXCEDENTE_DESDE",
    "unidad": "UNIDAD",
    "fuente": "FUENTE",
    "origen": "ORIGEN",
}

NUMERIC_COLS = {"VALOR", "DESDE", "HASTA", "MONTO_FIJO", "EXCEDENTE_DESDE", "PORCENTAJE"}


def update_excel(excel_path: Path, rebuild: bool = False):
    json_path = NORMALIZED_DIR / "parametros_arca.json"
    data = json.loads(json_path.read_text(encoding="utf-8"))

    wb = load_workbook(excel_path)
    ws = wb[SHEET_NAME]

    if rebuild:
        ws.delete_rows(2, ws.max_row)

    headers = [c.value for c in ws[1]]
    excel_cols = {h: i+1 for i, h in enumerate(headers)}

    existing = {}
    for r in range(2, ws.max_row + 1):
        key = (
            ws.cell(r, excel_cols["CONCEPTO"]).value,
            ws.cell(r, excel_cols["IMPUESTO"]).value,
            ws.cell(r, excel_cols["AÑO"]).value,
        )
        if all(key):
            existing[key] = r

    for item in data:
        key = (item["concepto"], item["impuesto"], item["anio"])

        row = existing.get(key, ws.max_row + 1)
        existing[key] = row

        for j_key, x_key in HEADER_MAP.items():
            if j_key in item:
                cell = ws.cell(row, excel_cols[x_key])
                value = item[j_key]
                cell.value = value

                if x_key in NUMERIC_COLS and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

    wb.save(excel_path)
    print(f"✅ Excel actualizado: {excel_path}")

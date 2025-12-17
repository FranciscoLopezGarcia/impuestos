# parametros_arca.py
import json
from openpyxl import load_workbook
from pathlib import Path


def update_excel(json_path: Path, excel_path: Path):
    SHEET_NAME = "Parametros_ARCA"

    with json_path.open(encoding="utf-8") as f:
        parametros = json.load(f)

    wb = load_workbook(excel_path)

    headers = [
        "concepto",
        "impuesto",
        "anio",
        "valor",
        "desde",
        "hasta",
        "monto_fijo",
        "porcentaje",
        "excedente_desde",
        "unidad",
        "fuente",
        "origen",
    ]

    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(headers)
    else:
        ws = wb[SHEET_NAME]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

    for p in parametros:
        es_escala = any(
            p.get(k) is not None
            for k in ("desde", "hasta", "porcentaje", "monto_fijo")
        )

        ws.append([
            p.get("concepto"),
            p.get("impuesto"),
            p.get("anio"),
            None if es_escala else p.get("valor_num"),
            p.get("desde"),
            p.get("hasta"),
            p.get("monto_fijo"),
            p.get("porcentaje"),
            p.get("excedente_desde"),
            p.get("unidad"),
            p.get("fuente"),
            p.get("origen"),
        ])

    try:
        wb.save(excel_path)
        print("✅ Excel actualizado correctamente")
    except PermissionError:
        print("❌ Cerrá el Excel antes de ejecutar el script")

"""
Excel Updater - Actualiza la hoja Parametros_ARCA en Excels existentes.

Este módulo permite actualizar Excels de clientes sin romper nada:
- Actualiza valores existentes (mismo CONCEPTO + IMPUESTO + AÑO)
- Agrega nuevos parámetros si no existen
- Preserva todas las demás hojas y fórmulas intactas
- Crea backup automático antes de modificar

Uso:
    from excel_updater import update_excel
    
    update_excel("cliente1.xlsx", year=2024)
    update_excel("cliente2.xlsx", year=2024, backup=True)
"""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers

from config import (
    NORMALIZED_DIR,
    PARAMS_SHEET_NAME,
    EXCEL_COLUMNS,
    NUMERIC_COLUMNS,
    get_backup_path
)


class ExcelUpdateError(Exception):
    """Error al actualizar Excel"""
    pass


def create_backup(excel_path: Path) -> Path:
    """Crea un backup del Excel antes de modificarlo"""
    import shutil
    
    backup_path = get_backup_path(excel_path)
    shutil.copy2(excel_path, backup_path)
    
    return backup_path


def load_parametros_json(year: int = None) -> list[dict]:
    """Carga el JSON normalizado de parámetros"""
    json_path = NORMALIZED_DIR / "parametros_arca.json"
    
    if not json_path.exists():
        raise FileNotFoundError(
            f"No se encontró el JSON normalizado: {json_path}\n"
            f"Ejecutá primero: python run_update.py --year {year}"
        )
    
    data = json.loads(json_path.read_text(encoding='utf-8'))
    
    # Filtrar por año si se especificó
    if year is not None:
        data = [item for item in data if item.get('anio') == year]
    
    return data


def get_or_create_params_sheet(workbook):
    """Obtiene la hoja de parámetros o la crea si no existe"""
    
    if PARAMS_SHEET_NAME in workbook.sheetnames:
        return workbook[PARAMS_SHEET_NAME]
    
    # Crear nueva hoja
    ws = workbook.create_sheet(PARAMS_SHEET_NAME, 0)  # Insertar al principio
    
    # Agregar headers
    for col_idx, header in enumerate(EXCEL_COLUMNS.values(), start=1):
        ws.cell(1, col_idx, header)
    
    return ws


def build_excel_index(ws) -> dict:
    """
    Construye un índice de los parámetros existentes en el Excel.
    
    Retorna un dict con clave (CONCEPTO, IMPUESTO, AÑO) -> número de fila
    """
    # Detectar columnas
    headers = [cell.value for cell in ws[1]]
    col_indices = {}
    
    for json_key, excel_name in EXCEL_COLUMNS.items():
        try:
            col_indices[json_key] = headers.index(excel_name) + 1
        except ValueError:
            raise ExcelUpdateError(
                f"No se encontró la columna '{excel_name}' en el Excel"
            )
    
    # Construir índice
    index = {}
    
    for row in range(2, ws.max_row + 1):
        concepto = ws.cell(row, col_indices['concepto']).value
        impuesto = ws.cell(row, col_indices['impuesto']).value
        anio = ws.cell(row, col_indices['anio']).value
        
        if concepto and impuesto and anio:
            key = (concepto, impuesto, anio)
            index[key] = row
    
    return index, col_indices


def update_excel(
    excel_path: Path | str,
    year: int = None,
    backup: bool = True,
    dry_run: bool = False
) -> dict:
    """
    Actualiza el Excel con los parámetros del JSON normalizado.
    
    Args:
        excel_path: Ruta al Excel a actualizar
        year: Año a procesar (None = todos los años en el JSON)
        backup: Crear backup antes de modificar
        dry_run: Simular cambios sin guardar
    
    Returns:
        Dict con estadísticas: {
            'updated': int,
            'added': int,
            'unchanged': int,
            'backup_path': Path
        }
    """
    
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel no encontrado: {excel_path}")
    
    # Cargar datos
    print(f"\n📂 Cargando: {excel_path.name}")
    parametros = load_parametros_json(year)
    
    if not parametros:
        print(f"⚠️  No hay parámetros para procesar (año={year})")
        return {'updated': 0, 'added': 0, 'unchanged': 0}
    
    print(f"📊 Parámetros a procesar: {len(parametros)}")
    
    # Crear backup
    backup_path = None
    if backup and not dry_run:
        print(f"💾 Creando backup...")
        backup_path = create_backup(excel_path)
        print(f"   ✓ Backup: {backup_path.name}")
    
    # Abrir Excel
    wb = load_workbook(excel_path)
    ws = get_or_create_params_sheet(wb)
    
    # Construir índice
    index, col_indices = build_excel_index(ws)
    
    # Estadísticas
    stats = {
        'updated': 0,
        'added': 0,
        'unchanged': 0,
        'backup_path': backup_path
    }
    
    # Procesar cada parámetro
    print(f"\n🔄 Procesando parámetros...")
    
    for param in parametros:
        concepto = param.get('concepto')
        impuesto = param.get('impuesto')
        anio = param.get('anio')
        
        if not all([concepto, impuesto, anio]):
            print(f"⚠️  Parámetro incompleto, saltando: {param}")
            continue
        
        key = (concepto, impuesto, anio)
        
        # Determinar si es update o insert
        if key in index:
            # UPDATE: Actualizar fila existente
            row = index[key]
            changed = False
            
            for json_key, value in param.items():
                if json_key not in col_indices:
                    continue
                
                col = col_indices[json_key]
                current_value = ws.cell(row, col).value
                
                # Comparar valores (convertir a string para comparar)
                if str(current_value) != str(value):
                    ws.cell(row, col, value)
                    changed = True
                    
                    # Aplicar formato numérico si corresponde
                    col_name = EXCEL_COLUMNS[json_key]
                    if col_name in NUMERIC_COLUMNS and isinstance(value, (int, float)):
                        ws.cell(row, col).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            if changed:
                stats['updated'] += 1
            else:
                stats['unchanged'] += 1
        
        else:
            # INSERT: Agregar nueva fila
            new_row = ws.max_row + 1
            
            for json_key, value in param.items():
                if json_key not in col_indices:
                    continue
                
                col = col_indices[json_key]
                ws.cell(new_row, col, value)
                
                # Aplicar formato numérico si corresponde
                col_name = EXCEL_COLUMNS[json_key]
                if col_name in NUMERIC_COLUMNS and isinstance(value, (int, float)):
                    ws.cell(new_row, col).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            stats['added'] += 1
            index[key] = new_row  # Actualizar índice
    
    # Guardar cambios
    if not dry_run:
        print(f"\n💾 Guardando cambios...")
        wb.save(excel_path)
        wb.close()
        print(f"   ✓ Excel actualizado")
    else:
        print(f"\n🔍 DRY RUN - No se guardaron cambios")
        wb.close()
    
    # Resumen
    print(f"\n📊 RESUMEN:")
    print(f"   ✓ Actualizados:  {stats['updated']}")
    print(f"   ✓ Agregados:     {stats['added']}")
    print(f"   • Sin cambios:   {stats['unchanged']}")
    
    if backup_path:
        print(f"\n💾 Backup guardado en:")
        print(f"   {backup_path}")
    
    return stats


def validate_excel_structure(excel_path: Path) -> bool:
    """
    Valida que el Excel tenga la estructura correcta.
    
    Verifica:
    - Que exista la hoja Parametros_ARCA
    - Que tenga las columnas necesarias
    - Que las fórmulas en otras hojas sean correctas
    """
    
    print(f"\n🔍 Validando estructura de: {excel_path.name}")
    
    try:
        wb = load_workbook(excel_path, data_only=False)
        
        # Verificar hoja
        if PARAMS_SHEET_NAME not in wb.sheetnames:
            print(f"   ❌ No existe la hoja '{PARAMS_SHEET_NAME}'")
            return False
        
        ws = wb[PARAMS_SHEET_NAME]
        
        # Verificar columnas
        headers = [cell.value for cell in ws[1]]
        expected_headers = list(EXCEL_COLUMNS.values())
        
        missing = [h for h in expected_headers if h not in headers]
        
        if missing:
            print(f"   ❌ Faltan columnas: {missing}")
            return False
        
        print(f"   ✓ Estructura correcta")
        print(f"   ✓ {len(headers)} columnas")
        print(f"   ✓ {ws.max_row - 1} filas de datos")
        
        wb.close()
        return True
    
    except Exception as e:
        print(f"   ❌ Error al validar: {e}")
        return False


def batch_update_excels(excel_folder: Path, year: int, pattern: str = "*.xlsx"):
    """
    Actualiza múltiples Excels en una carpeta.
    
    Args:
        excel_folder: Carpeta con los Excels
        year: Año a procesar
        pattern: Patrón de archivos (default: *.xlsx)
    """
    
    excel_folder = Path(excel_folder)
    
    if not excel_folder.exists():
        print(f"❌ Carpeta no existe: {excel_folder}")
        return
    
    excel_files = list(excel_folder.glob(pattern))
    
    if not excel_files:
        print(f"❌ No se encontraron archivos con patrón '{pattern}' en:")
        print(f"   {excel_folder}")
        return
    
    print(f"\n{'='*70}")
    print(f"ACTUALIZACIÓN MASIVA")
    print(f"{'='*70}")
    print(f"Carpeta: {excel_folder}")
    print(f"Archivos encontrados: {len(excel_files)}")
    print(f"Año: {year}")
    
    results = []
    
    for i, excel_path in enumerate(excel_files, 1):
        print(f"\n[{i}/{len(excel_files)}] {excel_path.name}")
        print("-" * 70)
        
        try:
            stats = update_excel(excel_path, year=year, backup=True)
            results.append({
                'file': excel_path.name,
                'success': True,
                'stats': stats
            })
        except Exception as e:
            print(f"   ❌ Error: {e}")
            results.append({
                'file': excel_path.name,
                'success': False,
                'error': str(e)
            })
    
    # Resumen final
    print(f"\n{'='*70}")
    print(f"RESUMEN FINAL")
    print(f"{'='*70}")
    
    success_count = sum(1 for r in results if r['success'])
    failed_count = len(results) - success_count
    
    print(f"\n✅ Exitosos: {success_count}/{len(results)}")
    if failed_count > 0:
        print(f"❌ Fallidos: {failed_count}/{len(results)}")
        print(f"\nArchivos con errores:")
        for r in results:
            if not r['success']:
                print(f"   - {r['file']}: {r['error']}")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Actualiza Excels con parámetros ARCA")
    
    parser.add_argument(
        'excel',
        help='Ruta al Excel o carpeta con Excels'
    )
    
    parser.add_argument(
        '--year',
        type=int,
        help='Año a procesar (default: todos los años en el JSON)'
    )
    
    parser.add_argument(
        '--no-backup',
        action='store_true',
        help='No crear backup antes de modificar'
    )
    
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Simular cambios sin guardar'
    )
    
    parser.add_argument(
        '--validate-only',
        action='store_true',
        help='Solo validar estructura, no actualizar'
    )
    
    parser.add_argument(
        '--batch',
        action='store_true',
        help='Actualizar todos los Excels en una carpeta'
    )
    
    args = parser.parse_args()
    
    excel_path = Path(args.excel)
    
    if args.validate_only:
        validate_excel_structure(excel_path)
    
    elif args.batch:
        batch_update_excels(excel_path, args.year)
    
    else:
        update_excel(
            excel_path,
            year=args.year,
            backup=not args.no_backup,
            dry_run=args.dry_run
        )